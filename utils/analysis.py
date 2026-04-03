"""
Analysis runner and Excel export for Octagonal Footing Design.
Handles load combination iteration, max ratio tracking, and output generation.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from utils.calculations import OctagonFootingCalc


def parse_reactions(raw_text, use_end_node=True):
    """
    Parse STAAD Pro reaction data from pasted text.
    
    Expected format (tab or space separated):
        Beam  L/C  Node  Fx(kN)  Fy(kN)  Fz(kN)  Mx(kN-m)  My(kN-m)  Mz(kN-m)
    
    The reactions typically come in pairs (node 1 and node 2).
    We want the 'end' node (node 2) by default, or 'start' node (node 1).
    
    Args:
        raw_text: pasted text from STAAD Pro
        use_end_node: True for End node (typically node 2), False for Start
        
    Returns:
        List of dicts with keys: beam, lc, lc_name, node, Fx, Fy, Fz, Mx, My, Mz
    """
    lines = raw_text.strip().split('\n')
    data = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Skip header lines
        parts = line.split()
        if len(parts) < 9:
            # Try tab-separated
            parts = line.split('\t')
        if len(parts) < 9:
            continue
            
        try:
            beam = int(parts[0])
            # L/C might be like "101 LC101" or just "101"
            # Handle multi-part LC name
            # Try parsing: first col is beam, then LC info, then node, then 6 numeric values
            # Find the pattern: int, string(LC), int(node), float*6
            
            # Strategy: try from the end - last 6 are floats, before that is node (int)
            floats = []
            idx = len(parts) - 1
            while idx >= 0 and len(floats) < 6:
                try:
                    floats.insert(0, float(parts[idx]))
                    idx -= 1
                except ValueError:
                    break
            
            if len(floats) != 6:
                continue
            
            node = int(parts[idx])
            idx -= 1
            
            # Everything between beam and node is the LC name
            lc_parts = parts[1:idx + 1]
            lc_name = ' '.join(lc_parts)
            
            # Extract LC number from name
            lc_num = None
            for p in lc_parts:
                try:
                    lc_num = int(p)
                    break
                except ValueError:
                    # Try extracting number from string like "LC101"
                    import re
                    match = re.search(r'LC(\d+)', p, re.IGNORECASE)
                    if match:
                        lc_num = int(match.group(1))
                        break
                    match = re.search(r'(\d+)', p)
                    if match:
                        lc_num = int(match.group(1))
                        break
            
            if lc_num is None:
                lc_num = 0
            
            row = {
                'beam': beam,
                'lc': lc_num,
                'lc_name': lc_name,
                'node': node,
                'Fx': floats[0],
                'Fy': floats[1],
                'Fz': floats[2],
                'Mx': floats[3],
                'My': floats[4],
                'Mz': floats[5],
            }
            data.append(row)
        except (ValueError, IndexError):
            continue
    
    # Filter by node type
    if not data:
        return []
    
    # Group by LC to identify start/end nodes
    # For each LC, we take the row based on use_end_node
    from collections import defaultdict
    lc_groups = defaultdict(list)
    for row in data:
        lc_groups[row['lc']].append(row)
    
    filtered = []
    for lc, rows in lc_groups.items():
        if len(rows) >= 2:
            # Sort by node number
            rows.sort(key=lambda x: x['node'])
            if use_end_node:
                filtered.append(rows[-1])  # End node (higher node number)
            else:
                filtered.append(rows[0])   # Start node (lower node number)
        else:
            filtered.append(rows[0])
    
    return filtered


def compute_load_from_reaction(row):
    """
    Convert STAAD reaction to footing loads.
    P = -Fx (sign change for reaction → applied load)
    H = sqrt(Fy² + Fz²)
    M = sqrt(My² + Mz²)
    """
    P = -row['Fx']
    H = (row['Fy'] ** 2 + row['Fz'] ** 2) ** 0.5
    M = (row['My'] ** 2 + row['Mz'] ** 2) ** 0.5
    return P, H, M


def run_analysis(input_params, reactions, ds_mapping, use_end_node=True, progress_callback=None):
    """
    Run full analysis across all load combinations.
    
    Args:
        input_params: dict with keys Df, Tf, Dp, hp, gamma_c, Ds, gamma_s, Kp, mu, Q, q_allow
        reactions: list of reaction dicts from parse_reactions
        ds_mapping: dict mapping LC number → Ds value
        use_end_node: whether to use end or start node
        progress_callback: optional callback(current, total, message)
    
    Returns:
        dict with:
            'controlling_results': full results dict for the controlling LC
            'controlling_lc': LC name/number
            'max_ratios': {Ratio_OT, Ratio_SLD, Ratio_SBC, Ratio_max}
            'all_lc_ratios': list of (LC, Ratio_max, Ratio_OT, Ratio_SLD, Ratio_SBC)
    """
    # Create calculator
    calc = OctagonFootingCalc(
        Df=input_params['Df'],
        Tf=input_params['Tf'],
        Dp=input_params['Dp'],
        hp=input_params['hp'],
        gamma_c=input_params['gamma_c'],
        Ds=input_params.get('Ds', 0),
        gamma_s=input_params['gamma_s'],
        Kp=input_params['Kp'],
        mu=input_params['mu'],
        Q=input_params['Q'],
        q_allow=input_params['q_allow']
    )
    
    max_ratio_ot = 0
    max_ratio_sld = 0
    max_ratio_sbc = 0
    max_ratio_overall = 0
    controlling_lc = None
    controlling_results = None
    controlling_ratios = None
    all_lc_ratios = []
    
    total = len(reactions)
    for i, rxn in enumerate(reactions):
        lc = rxn['lc']
        lc_name = rxn.get('lc_name', str(lc))
        
        if progress_callback:
            progress_callback(i + 1, total, f"LC {lc_name}")
        
        # Get Ds for this LC
        Ds = ds_mapping.get(lc, input_params.get('Ds', 0))
        
        # Convert reaction to loads
        P, H, M = compute_load_from_reaction(rxn)
        
        # Run calculation
        result = calc.compute_ratios(P, H, M, Ds=Ds)
        
        all_lc_ratios.append({
            'LC': lc,
            'LC_name': lc_name,
            'P': P,
            'H': H,
            'M': M,
            'Ds': Ds,
            'Ratio_OT': result['Ratio_OT'],
            'Ratio_SLD': result['Ratio_SLD'],
            'Ratio_SBC': result['Ratio_SBC'],
            'Ratio_max': result['Ratio_max'],
        })
        
        # Update max ratios
        current_max = result['Ratio_max']
        
        if current_max > max_ratio_overall:
            max_ratio_overall = current_max
            controlling_lc = lc_name
            controlling_results = result['results']
            controlling_ratios = result
        
        # Also track individual max ratios
        if result['Ratio_OT'] > max_ratio_ot:
            max_ratio_ot = result['Ratio_OT']
        if result['Ratio_SLD'] > max_ratio_sld:
            max_ratio_sld = result['Ratio_SLD']
        if result['Ratio_SBC'] > max_ratio_sbc:
            max_ratio_sbc = result['Ratio_SBC']
    
    return {
        'controlling_results': controlling_results,
        'controlling_lc': controlling_lc,
        'controlling_ratios': controlling_ratios,
        'max_ratios': {
            'Ratio_OT': max_ratio_ot,
            'Ratio_SLD': max_ratio_sld,
            'Ratio_SBC': max_ratio_sbc,
            'Ratio_max': max_ratio_overall,
        },
        'all_lc_ratios': all_lc_ratios,
    }


def export_analysis_xlsx(analysis_result, filepath, job_info=None):
    """
    Export analysis results to analysis.xlsx.
    
    Format: Column A = Parameter name, Column B = Value, Column C = Unit
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis Results"
    
    # Styles
    header_font = Font(name='Calibri', bold=True, size=12)
    section_font = Font(name='Calibri', bold=True, size=11, color='003366')
    normal_font = Font(name='Calibri', size=10)
    value_font = Font(name='Calibri', size=10, bold=True)
    
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_font_white = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    section_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    result_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    
    row = 1
    
    def write_header(text):
        nonlocal row
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).fill = header_fill
        row += 1
    
    def write_section(text):
        nonlocal row
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = section_font
        cell.fill = section_fill
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).fill = section_fill
        row += 1
    
    def write_row(name, value, unit=''):
        nonlocal row
        cell_a = ws.cell(row=row, column=1, value=name)
        cell_a.font = normal_font
        cell_a.border = thin_border
        
        cell_b = ws.cell(row=row, column=2)
        if isinstance(value, float):
            cell_b.value = round(value, 4)
        else:
            cell_b.value = value
        cell_b.font = value_font
        cell_b.alignment = Alignment(horizontal='center')
        cell_b.border = thin_border
        
        cell_c = ws.cell(row=row, column=3, value=unit)
        cell_c.font = normal_font
        cell_c.alignment = Alignment(horizontal='center')
        cell_c.border = thin_border
        row += 1
    
    def write_result_row(name, value, unit='', status=''):
        nonlocal row
        cell_a = ws.cell(row=row, column=1, value=name)
        cell_a.font = normal_font
        cell_a.fill = result_fill
        cell_a.border = thin_border
        
        cell_b = ws.cell(row=row, column=2)
        if isinstance(value, float):
            cell_b.value = round(value, 4)
        else:
            cell_b.value = value
        cell_b.font = value_font
        cell_b.alignment = Alignment(horizontal='center')
        cell_b.fill = result_fill
        cell_b.border = thin_border
        
        cell_c = ws.cell(row=row, column=3, value=unit if not status else status)
        cell_c.font = Font(name='Calibri', size=10, bold=True, 
                          color='008000' if status == 'OK' else ('FF0000' if status == 'NG' else '000000'))
        cell_c.alignment = Alignment(horizontal='center')
        cell_c.fill = result_fill
        cell_c.border = thin_border
        row += 1
    
    res = analysis_result.get('controlling_results', {})
    ratios = analysis_result.get('max_ratios', {})
    ctrl_lc = analysis_result.get('controlling_lc', 'N/A')
    
    if not res:
        ws.cell(row=1, column=1, value="No valid load combinations found.")
        wb.save(filepath)
        return
    
    # ---- Title ----
    write_header("OCTAGONAL SPREAD FOOTING ANALYSIS")
    row += 1
    
    # ---- Job Info ----
    if job_info:
        write_section("Job Information")
        if job_info.get('job_name'):
            write_row("Job Name", job_info['job_name'])
        if job_info.get('job_number'):
            write_row("Job Number", job_info['job_number'])
        if job_info.get('subject'):
            write_row("Subject", job_info['subject'])
        if job_info.get('originator'):
            write_row("Originator", job_info['originator'])
        if job_info.get('checker'):
            write_row("Checker", job_info['checker'])
        row += 1
    
    # ---- Input Data ----
    write_section("Input Data - Footing")
    write_row("Ftg. Base Length, Df", res.get('Df', 0), 'm')
    write_row("Ftg. Base Thickness, Tf", res.get('Tf', 0), 'm')
    write_row("Oct. Pier Length, Dp", res.get('Dp', 0), 'm')
    write_row("Oct. Pier Height, hp", res.get('hp', 0), 'm')
    write_row("Concrete Unit Wt., γc", res.get('gamma_c', 0), 'kN/m³')
    write_row("Soil Depth, Ds", res.get('Ds', 0), 'm')
    write_row("Soil Unit Wt., γs", res.get('gamma_s', 0), 'kN/m³')
    write_row("Pass. Press. Coef., Kp", res.get('Kp', 0))
    write_row("Coef. of Base Friction, μ", res.get('mu', 0))
    write_row("Uniform Surcharge, Q", res.get('Q', 0), 'kN/m²')
    write_row("SB Capacity, q_allow", res.get('q_allow', 0), 'kN/m²')
    row += 1
    
    # ---- Controlling Load Case ----
    write_section("Controlling Loading Data (LC: " + str(ctrl_lc) + ")")
    write_row("Applied Vert. Load, P", res.get('P', 0), 'kN')
    write_row("Applied Horiz. Load, H", res.get('H', 0), 'kN')
    write_row("Applied Moment, M", res.get('M', 0), 'kN-m')
    write_row("Load Case", ctrl_lc)
    row += 1
    
    # ---- Footing Base Properties ----
    write_section("Footing Base Properties")
    write_row("Dimension, Bf", res.get('Bf', 0), 'm')
    write_row("Footing Flat Side, Cf", res.get('Cf', 0), 'm')
    write_row("Footing Diagonal, Ef", res.get('Ef', 0), 'm')
    write_row("Footing Base Area, Af", res.get('Af', 0), 'm²')
    write_row("Footing Volume, Vf", res.get('Vf', 0), 'm³')
    write_row("Footing Inertia, If", res.get('If', 0), 'm⁴')
    row += 1
    
    # ---- Pier Properties ----
    write_section("Pier Properties")
    write_row("Dimension, Bp", res.get('Bp', 0), 'm')
    write_row("Pier Flat Side, Cp", res.get('Cp', 0), 'm')
    write_row("Pier Diagonal, Ep", res.get('Ep', 0), 'm')
    write_row("Pier Area, Ap", res.get('Ap', 0), 'm²')
    write_row("Pier Volume, Vp", res.get('Vp', 0), 'm³')
    row += 1
    
    # ---- Weights ----
    write_section("Pier, Surcharge, Soil, and Footing Base Weights")
    write_row("Pier Weight, Wp", res.get('Wp', 0), 'kN')
    write_row("Surcharge Load, Wq", res.get('Wq', 0), 'kN')
    write_row("Soil Weight, Ws", res.get('Ws', 0), 'kN')
    write_row("Ftg. Base Weight, Wf", res.get('Wf', 0), 'kN')
    row += 1
    
    # ---- Total Loads ----
    write_section("Total Resultant Load and Eccentricities")
    write_row("Total Vert. Load, ΣP", res.get('SP', 0), 'kN')
    write_row("Total Moment, ΣM", res.get('SM', 0), 'kN-m')
    write_row("Eccentricity, e", res.get('e', 0), 'm')
    write_row("Eccentricity Ratio, e/Df", res.get('e_Df', 0))
    row += 1
    
    # ---- Overturning Check ----
    write_section("Overturning Check")
    write_row("Mr", res.get('Mr', 0), 'kN-m')
    write_row("Mo", res.get('Mo', 0), 'kN-m')
    fs_ot = res.get('FS_ot', 'N.A.')
    if isinstance(fs_ot, (int, float)):
        write_row("FS(ot)", fs_ot)
    else:
        write_row("FS(ot)", fs_ot)
    row += 1
    
    # ---- Sliding Check ----
    write_section("Sliding Check")
    write_row("Passive Resist., PR", res.get('PR', 0), 'kN')
    write_row("Frict. Resist., FR", res.get('FR', 0), 'kN')
    fs_slid = res.get('FS_slid', 'N.A.')
    if isinstance(fs_slid, (int, float)):
        write_row("FS(slid)", fs_slid)
    else:
        write_row("FS(slid)", fs_slid)
    row += 1
    
    # ---- Bearing Pressure (Corners) ----
    write_section("Bearing Pressure (Axis through Corners)")
    write_row("Section Modulus, Sf", res.get('corners_Sf', 0), 'm³')
    write_row("Brg. Distance Coef., K", res.get('corners_K', 0))
    write_row("K*Df", res.get('corners_K_Df', 0), 'm')
    pct = res.get('corners_pct_brg_area', 0)
    write_row("%Brg. Area", pct, '%' if isinstance(pct, (int, float)) else '')
    write_row("Bearing Coef., L", res.get('corners_L', 'N.A.'))
    write_row("Gross Bearing, P(max)", res.get('corners_Pmax_gross', 0), 'kN/m²')
    write_row("Gross Bearing, P(min)", res.get('corners_Pmin_gross', 0), 'kN/m²')
    write_row("Net Press., Pmax(net)", res.get('corners_Pmax_net', 0), 'kN/m²')
    row += 1
    
    # ---- Bearing Pressure (Flat Sides) ----
    write_section("Bearing Pressure (Axis through Flat Sides)")
    write_row("Section Modulus, Sf", res.get('flat_Sf', 0), 'm³')
    write_row("Brg. Distance Coef., K", res.get('flat_K', 0))
    write_row("K*Df", res.get('flat_K_Df', 0), 'm')
    pct = res.get('flat_pct_brg_area', 0)
    write_row("%Brg. Area", pct, '%' if isinstance(pct, (int, float)) else '')
    write_row("Bearing Coef., L", res.get('flat_L', 'N.A.'))
    write_row("Gross Bearing, P(max)", res.get('flat_Pmax_gross', 0), 'kN/m²')
    write_row("Gross Bearing, P(min)", res.get('flat_Pmin_gross', 0), 'kN/m²')
    write_row("Net Press., Pmax(net)", res.get('flat_Pmax_net', 0), 'kN/m²')
    row += 1
    
    # ---- Summary of Results ----
    write_section("Summary of Results")
    fs_ot_val = res.get('FS_ot', 'N.A.')
    if isinstance(fs_ot_val, (int, float)):
        write_result_row("FS(ot)", round(fs_ot_val, 3))
    else:
        write_result_row("FS(ot)", fs_ot_val)
    
    fs_slid_val = res.get('FS_slid', 'N.A.')
    if isinstance(fs_slid_val, (int, float)):
        write_result_row("FS(slid)", round(fs_slid_val, 3))
    else:
        write_result_row("FS(slid)", fs_slid_val)
    
    pct_brg = res.get('pct_brg_area', 0)
    if isinstance(pct_brg, (int, float)):
        write_result_row("%Brg. Area", round(pct_brg, 2), '%')
    else:
        write_result_row("%Brg. Area", pct_brg)
    
    pmax = res.get('Pmax_gross', 0)
    if isinstance(pmax, (int, float)):
        write_result_row("Pmax(gross)", round(pmax, 3), 'kN/m²')
    else:
        write_result_row("Pmax(gross)", pmax)
    
    pnet = res.get('Pmax_net', 0)
    if isinstance(pnet, (int, float)):
        write_result_row("Pmax(net)", round(pnet, 3), 'kN/m²')
    else:
        write_result_row("Pmax(net)", pnet)
    row += 1
    
    # ---- Ratio Check ----
    write_section("Design Check Ratios (Max across all Load Combinations)")
    r_ot = ratios.get('Ratio_OT', 0)
    r_sld = ratios.get('Ratio_SLD', 0)
    r_sbc = ratios.get('Ratio_SBC', 0)
    r_max = ratios.get('Ratio_max', 0)
    
    write_result_row("Ratio Overturning", round(r_ot, 4), status='OK' if r_ot <= 1 else 'NG')
    write_result_row("Ratio Sliding", round(r_sld, 4), status='OK' if r_sld <= 1 else 'NG')
    write_result_row("Ratio Soil BC", round(r_sbc, 4), status='OK' if r_sbc <= 1 else 'NG')
    write_result_row("Ratio Max", round(r_max, 4), status='OK' if r_max <= 1 else 'NG')
    write_row("Controlling Load Case", ctrl_lc)
    row += 1
    
    # ---- All LC Ratios (Sheet 2) ----
    ws2 = wb.create_sheet("Load Combination Summary")
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 8
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 12
    ws2.column_dimensions['I'].width = 12
    ws2.column_dimensions['J'].width = 12
    
    headers = ['LC', 'LC Name', 'P (kN)', 'H (kN)', 'M (kN-m)', 'Ds (m)', 
               'Ratio OT', 'Ratio SLD', 'Ratio SBC', 'Ratio Max']
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    all_lc = analysis_result.get('all_lc_ratios', [])
    for i, lc_data in enumerate(all_lc, 2):
        ws2.cell(row=i, column=1, value=lc_data['LC']).border = thin_border
        ws2.cell(row=i, column=2, value=lc_data['LC_name']).border = thin_border
        ws2.cell(row=i, column=3, value=round(lc_data['P'], 3)).border = thin_border
        ws2.cell(row=i, column=4, value=round(lc_data['H'], 3)).border = thin_border
        ws2.cell(row=i, column=5, value=round(lc_data['M'], 3)).border = thin_border
        ws2.cell(row=i, column=6, value=round(lc_data['Ds'], 3)).border = thin_border
        ws2.cell(row=i, column=7, value=round(lc_data['Ratio_OT'], 4)).border = thin_border
        ws2.cell(row=i, column=8, value=round(lc_data['Ratio_SLD'], 4)).border = thin_border
        ws2.cell(row=i, column=9, value=round(lc_data['Ratio_SBC'], 4)).border = thin_border
        
        cell_max = ws2.cell(row=i, column=10, value=round(lc_data['Ratio_max'], 4))
        cell_max.border = thin_border
        if lc_data['Ratio_max'] > 1:
            cell_max.font = Font(color='FF0000', bold=True)
        else:
            cell_max.font = Font(color='008000')
    
    wb.save(filepath)
    return filepath
